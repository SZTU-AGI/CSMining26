#!/usr/bin/env python3
from __future__ import annotations
import argparse,json
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

VERSION="V6.6_DEADLINE_FINAL_BATCH_1"
CP=[f"CP{i}" for i in range(1,42)]; ALLOWED={"0","1","N/A"}

def load(p): return json.loads(Path(p).read_text(encoding='utf-8'))
def save(v,p): Path(p).write_text(json.dumps(v,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')

def re_number(case):
    for k in ('re_number_candidate','output_identifier','re_number','RE Number'):
        if case.get(k): return str(case[k])
    raise RuntimeError(f"No RE number for {case.get('case_uid')}")

def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--run-root',type=Path,required=True)
    ap.add_argument('--manifest',type=Path,default=Path('/home/MeggieYu/freca/core_v1/results_v2/logical_case_manifest_v1.json'))
    ap.add_argument('--template',type=Path,default=Path('/home/MeggieYu/freca/Task2/submission_template.xlsx'))
    ap.add_argument('--output',type=Path,default=None)
    a=ap.parse_args(); out=a.output or a.run_root/'submission_v6_6_deadline_final.xlsx'
    manifest=load(a.manifest); cases=sorted(manifest.get('cases') or [],key=lambda x:int(x['serial']))
    if len(cases)!=100: raise RuntimeError(f"Expected 100 cases, got {len(cases)}")
    grid={}; audit=[]; counts=Counter()
    for case in cases:
        uid=str(case['case_uid']); p=a.run_root/'cases'/uid/'final_case.json'
        if not p.exists(): raise RuntimeError(f"Missing {p}")
        d=load(p)
        if d.get('version')!=VERSION: raise RuntimeError(f"Wrong version {p}: {d.get('version')}")
        for cp in CP:
            row=(d.get('verdicts') or {}).get(cp) or {}; v=str(row.get('verdict') or '')
            if v not in ALLOWED: raise RuntimeError(f"Invalid {uid}/{cp}: {v!r}")
            grid[(uid,cp)]=v; counts[v]+=1
            audit.append({'case_uid':uid,'cp_id':cp,'verdict':v,'reason':row.get('reason'),'evidence_ids':row.get('evidence_ids') or [],'na_verified':row.get('na_verified',False)})
    wb=load_workbook(a.template); ws=wb.active
    header=[ws.cell(1,c).value for c in range(1,43)]
    if header!=['RE Number',*CP]: raise RuntimeError(f"Template header mismatch: {header}")
    # Official package template is header-only. Clear any stray rows to be safe without changing columns/order.
    if ws.max_row>1: ws.delete_rows(2,ws.max_row-1)
    for case in cases:
        uid=str(case['case_uid']); ws.append([re_number(case),*[grid[(uid,cp)] for cp in CP]])
    out.parent.mkdir(parents=True,exist_ok=True); wb.save(out)
    # Re-read validation.
    wb2=load_workbook(out,data_only=True); ws2=wb2.active
    if ws2.max_row!=101 or ws2.max_column!=42: raise RuntimeError(f"Output shape invalid {ws2.max_row}x{ws2.max_column}")
    vals=Counter(str(ws2.cell(r,c).value) for r in range(2,102) for c in range(2,43))
    if set(vals)-ALLOWED: raise RuntimeError(f"Unexpected output labels: {vals}")
    audit_path=a.run_root/'submission_v6_6_deadline_final_audit.json'
    save({'version':VERSION,'counts':dict(counts),'output':str(out),'rows':audit},audit_path)
    print('FINAL SUBMISSION READY')
    print('Output:',out); print('Counts:',dict(counts)); print('Audit:',audit_path)
if __name__=='__main__': main()
