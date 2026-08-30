#!/usr/bin/env python3
from __future__ import annotations

"""Assemble the official 100 x 41 FRECA Task2 submission from V6.5 decisions."""

import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

CP_IDS = [f"CP{i}" for i in range(1, 42)]
HEADER = ["RE Number", *CP_IDS]
ALLOWED = {"0", "1", "N/A"}


def load(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def save(value, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def case_re_number(case: dict, decision_grid: dict, case_uid: str) -> str:
    for key in ("re_number_candidate", "output_identifier", "re_number", "RE Number"):
        value = case.get(key)
        if value:
            return str(value)
    for cp in CP_IDS:
        d = decision_grid.get((case_uid, cp))
        if d and d.get("output_identifier"):
            return str(d["output_identifier"])
    raise RuntimeError(f"Cannot resolve RE Number for {case_uid}")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-root", type=Path, required=True)
    ap.add_argument("--manifest", type=Path,
                    default=Path("/home/MeggieYu/freca/core_v1/results_v2/logical_case_manifest_v1.json"))
    ap.add_argument("--template", type=Path,
                    default=Path("/home/MeggieYu/freca/Task2/submission_template.xlsx"))
    ap.add_argument("--output", type=Path, default=None)
    ap.add_argument("--audit", type=Path, default=None)
    args = ap.parse_args()

    output = args.output or (args.run_root / "submission_v6_5_final.xlsx")
    audit_path = args.audit or (args.run_root / "submission_v6_5_final_audit.json")
    manifest = load(args.manifest)
    cases = sorted(manifest.get("cases") or [], key=lambda x: int(x["serial"]))
    if len(cases) != 100 or [int(x["serial"]) for x in cases] != list(range(1, 101)):
        raise RuntimeError("Manifest is not exactly serial 1..100")

    decisions = {}
    duplicates = []
    for p in sorted(args.run_root.glob("worker-*/tasks/case-*/CP*/decision.json")):
        d = load(p)
        key = (str(d.get("case_uid")), str(d.get("cp_id")))
        if key in decisions:
            duplicates.append({"key": key, "first": decisions[key].get("_path"), "second": str(p)})
        d["_path"] = str(p)
        decisions[key] = d
    if duplicates:
        raise RuntimeError(f"Duplicate decisions found: {duplicates[:3]}")

    expected = {(f"case-{i:03d}", cp) for i in range(1,101) for cp in CP_IDS}
    missing = sorted(expected - set(decisions))
    extra = sorted(set(decisions) - expected)
    if missing or extra:
        raise RuntimeError(f"Decision grid incomplete: missing={len(missing)} extra={len(extra)}; first_missing={missing[:5]}")

    wb = load_workbook(args.template)
    ws = wb.active
    existing_header = [ws.cell(1, c).value for c in range(1, 43)]
    if existing_header != HEADER:
        raise RuntimeError(f"Official template header mismatch: {existing_header}")
    if ws.max_row != 1:
        raise RuntimeError(f"Expected header-only official template, found {ws.max_row} rows")

    labels = Counter()
    outcomes = Counter()
    applicability = Counter()
    na_coordinates = []
    fallback_coordinates = []

    for case in cases:
        uid = str(case["case_uid"])
        row = [case_re_number(case, decisions, uid)]
        for cp in CP_IDS:
            d = decisions[(uid, cp)]
            label = str(d.get("fold_label") or "")
            if label not in ALLOWED:
                raise RuntimeError(f"{uid}/{cp} invalid label {label!r}")
            row.append(label)
            labels[label] += 1
            outcome = str(d.get("common_internal_outcome") or "")
            outcomes[outcome] += 1
            fold = d.get("fold_decision") or {}
            app_decision = str(fold.get("v6_5_applicability_decision") or "")
            if app_decision:
                applicability[app_decision] += 1
            if label == "N/A":
                na_coordinates.append({
                    "case_uid": uid,
                    "cp_id": cp,
                    "re_number": row[0],
                    "internal_outcome": outcome,
                    "finality": d.get("fold_finality"),
                    "applicability_decision": app_decision,
                })
            if d.get("benchmark_fallback"):
                fallback_coordinates.append({"case_uid": uid, "cp_id": cp, "label": label, "outcome": outcome})
        ws.append(row)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    # Re-read the exact graded artefact.
    check = load_workbook(output, data_only=True).active
    if check.max_row != 101 or check.max_column != 42:
        raise RuntimeError(f"Written submission shape {check.max_row}x{check.max_column}, expected 101x42")
    if [check.cell(1,c).value for c in range(1,43)] != HEADER:
        raise RuntimeError("Written submission header changed")
    reread = Counter()
    for r in range(2,102):
        if not str(check.cell(r,1).value or "").strip():
            raise RuntimeError(f"Submission row {r} has empty RE Number")
        for c in range(2,43):
            value = str(check.cell(r,c).value or "").strip()
            if value not in ALLOWED:
                raise RuntimeError(f"Invalid submitted value at row={r} col={c}: {value!r}")
            reread[value] += 1
    if sum(reread.values()) != 4100:
        raise RuntimeError("Submission does not contain 4100 verdict cells")

    audit = {
        "schema": "freca-v6.5-final-submission-audit-v1",
        "semantic_version": "V6.5_FINAL_0_1_NA",
        "run_root": str(args.run_root),
        "template": str(args.template),
        "output": str(output),
        "coordinate_count": 4100,
        "label_counts": dict(sorted(reread.items())),
        "internal_outcome_counts": dict(sorted(outcomes.items())),
        "applicability_decision_counts": dict(sorted(applicability.items())),
        "na_count": len(na_coordinates),
        "na_coordinates": na_coordinates,
        "benchmark_fallback_count": len(fallback_coordinates),
        "benchmark_fallback_coordinates": fallback_coordinates,
        "row_order": "logical_case_manifest serial ascending 1..100",
        "allowed_values": ["0", "1", "N/A"],
        "submission_verified_by_reread": True,
    }
    save(audit, audit_path)

    print("# FRECA V6.5 final submission")
    print("Coordinates: 4100")
    print("Labels:", dict(sorted(reread.items())))
    print("N/A coordinates:", len(na_coordinates))
    print("Output:", output)
    print("Audit:", audit_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
