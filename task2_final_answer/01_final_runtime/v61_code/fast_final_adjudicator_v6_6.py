#!/usr/bin/env python3
from __future__ import annotations

import argparse, hashlib, json, os, re, sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Runtime imports are from the installed FRECA core.
import freca_core_v1 as core
import production_runner_v1 as base

VERSION = "V6.6_DEADLINE_FINAL_BATCH_1"
ELEMENTS = {
    "Element-1": [f"CP{i}" for i in range(1, 8)],
    "Element-2": [f"CP{i}" for i in range(8, 17)],
    "Element-3": [f"CP{i}" for i in range(17, 29)],
    "Element-4": [f"CP{i}" for i in range(29, 42)],
}
ALLOWED = {"0", "1", "N/A"}


def load_json(p: Path):
    return json.loads(p.read_text(encoding="utf-8"))


def save_json(v: Any, p: Path):
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(json.dumps(v, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp.replace(p)


def sha(v: Any) -> str:
    s = json.dumps(v, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def load_cp_sheet(path: Path) -> tuple[dict[str,str], dict[str,str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headings = {}
    texts = {}
    for c in range(1, 42):
        cp = str(ws.cell(4, c).value or "").strip()
        txt = norm(ws.cell(3, c).value)
        # Carry the most recent non-empty subgroup heading from row 2.
        h = ws.cell(2, c).value
        if h:
            current_h = norm(h)
        headings[cp] = current_h if 'current_h' in locals() else ""
        texts[cp] = txt
    expected = {f"CP{i}" for i in range(1,42)}
    if set(texts) != expected:
        raise RuntimeError(f"Checking-point workbook malformed: got {len(texts)} CPs")
    return texts, headings


def load_policy_passages(pdf_path: Path, cache_path: Path) -> list[dict]:
    if cache_path.exists():
        v = load_json(cache_path)
        if isinstance(v, list) and v:
            return v
    try:
        from pypdf import PdfReader
    except Exception as e:
        raise RuntimeError("pypdf is required for policy retrieval") from e
    reader = PdfReader(str(pdf_path))
    passages = []
    for pageno, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        blocks = [norm(x) for x in re.split(r"\n\s*\n|(?<=\.)\s+(?=[A-Z][A-Za-z ]{2,40}:)", text) if norm(x)]
        seq = 0
        for block in blocks:
            # Chunk very long blocks without inventing text.
            for start in range(0, len(block), 1400):
                frag = block[start:start+1600].strip()
                if len(frag) < 35:
                    continue
                seq += 1
                passages.append({"id": f"RULES:p{pageno}:b{seq}", "text": frag, "page": pageno})
    if not passages:
        raise RuntimeError(f"No text extracted from policy PDF: {pdf_path}")
    save_json(passages, cache_path)
    return passages


def chunk_id(ch: dict) -> str:
    return str(ch.get("id") or ch.get("evidence_id") or ch.get("chunk_id") or "")


def chunk_text(ch: dict) -> str:
    return norm(ch.get("text") or ch.get("content") or ch.get("quote"))


def normalize_chunks(chunks: list[dict]) -> list[dict]:
    out=[]
    for ch in chunks:
        cid=chunk_id(ch); txt=chunk_text(ch)
        if cid and txt:
            x=dict(ch); x["id"]=cid; x["text"]=txt
            out.append(x)
    return out


def retrieve_policy(cp_text: str, policy_docs: list[dict], k: int=3) -> list[dict]:
    ranked = core.bm25_rank(cp_text, policy_docs, top_k=k)
    return [{"id": x["id"], "text": x["text"][:1800], "score": float(x.get("score",0))} for x in ranked if x.get("score",0) > 0]


def retrieve_evidence(query: str, chunks: list[dict], k: int=6) -> list[dict]:
    ranked = core.bm25_rank(query, chunks, top_k=k)
    out=[]
    for x in ranked:
        if x.get("score",0) <= 0:
            continue
        out.append({"id": x["id"], "text": x["text"][:1500], "score": float(x.get("score",0))})
    return out


def build_element_context(element: str, cp_texts: dict[str,str], headings: dict[str,str], policy_docs: list[dict], chunks: list[dict]):
    cps = ELEMENTS[element]
    cp_policy = {}
    evidence_by_cp = {}
    ev_union = {}
    policy_union = {}
    for cp in cps:
        p = retrieve_policy(cp_texts[cp], policy_docs, 3)
        cp_policy[cp] = p
        for row in p: policy_union[row["id"]] = row
        policy_query = " ".join(x["text"] for x in p)
        q = f"{headings.get(cp,'')} {cp_texts[cp]} {policy_query[:3500]}"
        ev = retrieve_evidence(q, chunks, 7)
        evidence_by_cp[cp] = ev
        for row in ev: ev_union[row["id"]] = row
    # Add a small track-diversity supplement: best lexical hit per track for the whole element.
    whole_query = " ".join(cp_texts[c] for c in cps)
    ranked = core.bm25_rank(whole_query, chunks, top_k=min(80, len(chunks)))
    seen_tracks=set()
    for x in ranked:
        cid=str(x.get("id") or "")
        m=re.match(r"([1-9])_", cid)
        tr=m.group(1) if m else None
        if tr and tr not in seen_tracks and x.get("score",0)>0:
            seen_tracks.add(tr)
            ev_union[cid] = {"id":cid,"text":x["text"][:1500],"score":float(x.get("score",0))}
        if len(seen_tracks)>=9: break
    return {
        "cps": cps,
        "cp_policy": cp_policy,
        "evidence_by_cp": evidence_by_cp,
        "policy_union": list(policy_union.values()),
        "evidence_union": list(ev_union.values()),
    }


def primary_prompts(element: str, ctx: dict, cp_texts: dict[str,str], headings: dict[str,str]):
    system = """You are an AI compliance auditor for the FRECA competition.
You receive ONLY: (a) checking-point text loaded verbatim from the official competition workbook, (b) policy excerpts automatically retrieved from the supplied Export Control Rules PDF, and (c) evidence excerpts automatically retrieved from the nine files of one farm case.
Do not use external facts and do not invent hidden checking-point rules.

For every supplied CP, output exactly one verdict:
- "1": the supplied evidence establishes that the checking point is satisfied.
- "0": the checking point applies but the evidence fails to establish compliance OR the evidence contradicts/fails the checking point. If there is material support and material contradiction, use 0.
- "N/A": the checking point itself does not apply to this farm's registered operations. N/A requires positive evidence of non-applicability; missing or ambiguous evidence is never enough.

Important:
- Judge each CP independently from its official text, retrieved policy, and grounded case evidence.
- A plan/procedure describing what should happen is not automatically proof that it actually happened when the CP asks about current/actual condition.
- Do not treat filenames as evidence.
- If applicability is uncertain, choose 0 rather than N/A.
- Return JSON only. Every CP must appear exactly once.
"""
    cp_lines=[]
    for cp in ctx["cps"]:
        pol = " | ".join(f"[{p['id']}] {p['text']}" for p in ctx["cp_policy"][cp])
        cp_lines.append(f"{cp} ({headings.get(cp,'')}): {cp_texts[cp]}\nAUTO-RETRIEVED POLICY FOR {cp}: {pol}")
    ev_lines=[f"[{x['id']}] {x['text']}" for x in ctx["evidence_union"]]
    user = f"""ELEMENT: {element}

OFFICIAL CHECKING POINTS AND AUTOMATIC POLICY RETRIEVAL:
{chr(10).join(cp_lines)}

AUTOMATICALLY RETRIEVED GROUNDED CASE EVIDENCE:
{chr(10).join(ev_lines)}

Return this exact JSON shape:
{{
  "verdicts": {{
    "CPx": {{"verdict":"1|0|N/A", "evidence_ids":["..."], "reason":"brief grounded reason"}}
  }}
}}
"""
    return system, user


def normalize_verdict(v: Any) -> str | None:
    s = str(v or "").strip().upper().replace(" ","")
    if s in {"1","COMPLIANT","TRUE"}: return "1"
    if s in {"0","NON-COMPLIANT","NONCOMPLIANT","FALSE"}: return "0"
    if s in {"N/A","NA","NOTAPPLICABLE","NOT_APPLICABLE"}: return "N/A"
    return None


def validate_primary(raw: dict, element: str, ctx: dict):
    expected=ELEMENTS[element]
    ev_index={x["id"]:x["text"] for x in ctx["evidence_union"]}
    rv = raw.get("verdicts") if isinstance(raw,dict) else None
    if not isinstance(rv,dict): rv={}
    out={}; errors=[]
    for cp in expected:
        row=rv.get(cp) if isinstance(rv.get(cp),dict) else {}
        verdict=normalize_verdict(row.get("verdict"))
        if verdict is None:
            verdict="0"; errors.append(f"{cp}:missing_or_invalid_verdict=>0")
        ids=[]
        for eid in row.get("evidence_ids") or []:
            eid=str(eid)
            if eid in ev_index and eid not in ids: ids.append(eid)
        # Compliance needs some grounded support. N/A will be independently verified below.
        if verdict=="1" and not ids:
            verdict="0"; errors.append(f"{cp}:ungrounded_1=>0")
        out[cp]={"verdict":verdict,"evidence_ids":ids,"reason":norm(row.get("reason"))[:1000]}
    return out, errors


def na_verify_prompt(na_rows: dict[str,dict], cp_texts: dict[str,str], ctx: dict):
    ev_index={x["id"]:x["text"] for x in ctx["evidence_union"]}
    blocks=[]
    for cp,row in na_rows.items():
        cited="\n".join(f"[{eid}] {ev_index[eid]}" for eid in row["evidence_ids"] if eid in ev_index)
        # Include CP-specific retrieved evidence too, not only primary citations.
        broader="\n".join(f"[{x['id']}] {x['text']}" for x in ctx["evidence_by_cp"].get(cp,[]))
        blocks.append(f"{cp}: {cp_texts[cp]}\nPRIMARY REASON: {row['reason']}\nCITED:\n{cited}\nBROADER CP EVIDENCE:\n{broader}")
    system="""You are an independent N/A verification gate. Verify only whether each checking point truly does NOT APPLY to this farm's registered operations.
N/A requires positive grounded evidence that the CP's triggering activity/entity/situation is absent, out of scope, or explicitly not applicable. Missing evidence is not enough. If the CP applies, might apply, or applicability is ambiguous, REJECT the N/A and output 0.
Do not judge ordinary compliance here. Return JSON only."""
    user=f"""Review these proposed N/A verdicts independently:
{chr(10).join(blocks)}

Return:
{{"na_verification":{{"CPx":{{"decision":"CONFIRM_NA|REJECT_TO_0","evidence_ids":["..."],"reason":"brief"}}}}}}
"""
    return system,user


def run_element(case_uid: str, element: str, chunks: list[dict], cp_texts, headings, policy_docs, out_dir: Path, model: str):
    out_path=out_dir/"elements"/f"{element}.json"
    ctx=build_element_context(element,cp_texts,headings,policy_docs,chunks)
    fingerprint=sha({"version":VERSION,"case":case_uid,"element":element,"model":model,
                     "cp":{c:cp_texts[c] for c in ELEMENTS[element]},
                     "policy":[(x['id'],x['text']) for x in ctx['policy_union']],
                     "evidence":[(x['id'],x['text']) for x in ctx['evidence_union']]})
    if out_path.exists():
        old=load_json(out_path)
        if old.get("input_fingerprint")==fingerprint and old.get("status")=="COMPLETE":
            return old, True
    sys_p,user_p=primary_prompts(element,ctx,cp_texts,headings)
    raw=core.deepseek_json(model=model,system_prompt=sys_p,user_prompt=user_p,thinking=False,max_tokens=5000)
    verdicts,errors=validate_primary(raw,element,ctx)
    na_rows={cp:r for cp,r in verdicts.items() if r['verdict']=="N/A"}
    na_raw=None
    if na_rows:
        ns,nu=na_verify_prompt(na_rows,cp_texts,ctx)
        na_raw=core.deepseek_json(model=model,system_prompt=ns,user_prompt=nu,thinking=False,max_tokens=2400)
        ver=(na_raw.get("na_verification") or {}) if isinstance(na_raw,dict) else {}
        for cp,row in na_rows.items():
            vr=ver.get(cp) if isinstance(ver.get(cp),dict) else {}
            if str(vr.get("decision") or "").strip().upper()!="CONFIRM_NA":
                verdicts[cp]["verdict"]="0"
                verdicts[cp]["na_rejected_reason"]=norm(vr.get("reason") or "independent N/A verifier did not confirm")[:1000]
            else:
                verdicts[cp]["na_verified"]=True
                verdicts[cp]["na_verification_reason"]=norm(vr.get("reason"))[:1000]
    result={
        "schema":"freca-v6.6-deadline-final-element-v1","version":VERSION,"status":"COMPLETE",
        "case_uid":case_uid,"element":element,"model":model,"input_fingerprint":fingerprint,
        "verdicts":verdicts,"validation_notes":errors,
        "retrieval":{"evidence_count":len(ctx['evidence_union']),"policy_count":len(ctx['policy_union'])},
        "prompt_log":{"system":sys_p,"user":user_p},
        "raw_model_response":raw,"raw_na_verification":na_raw,
    }
    save_json(result,out_path)
    return result,False


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--run-root",type=Path,required=True)
    ap.add_argument("--manifest",type=Path,default=Path("/home/MeggieYu/freca/core_v1/results_v2/logical_case_manifest_v1.json"))
    ap.add_argument("--cp-workbook",type=Path,default=Path("/home/MeggieYu/freca/Task2/checkingpoints_all_elements_onesheet.xlsx"))
    ap.add_argument("--policy-pdf",type=Path,default=Path("/home/MeggieYu/freca/Task2/1-Export Control (Plants and Plant Products)Rules 2021.pdf"))
    ap.add_argument("--case",action="append",default=[])
    ap.add_argument("--all",action="store_true")
    ap.add_argument("--model",default=os.environ.get("FRECA_ALIGNMENT_MODEL","deepseek-v4-flash"))
    args=ap.parse_args()

    manifest=load_json(args.manifest); base.validate_manifest(manifest); cmap=base.manifest_case_map(manifest)
    selected=[f"case-{i:03d}" for i in range(1,101)] if args.all else [base.normalize_case_selector(x) for x in args.case]
    if not selected: ap.error("Use --all or --case")
    selected=sorted(set(selected),key=lambda x:int(x.split('-')[1]))
    for c in selected:
        if c not in cmap: raise RuntimeError(f"Manifest missing {c}")
    cp_texts,headings=load_cp_sheet(args.cp_workbook)
    policy_docs=load_policy_passages(args.policy_pdf,args.run_root/"_cache"/"policy_passages.json")
    runtime=base.runtime_hashes(); parser_hash=runtime.get("freca_core_v1.py") or "unknown"
    case_root=Path(manifest["dataset_structure_profile"]["case_root"])
    args.run_root.mkdir(parents=True,exist_ok=True)
    print(f"V6.6 deadline-final batch adjudicator: cases={len(selected)} model={args.model}")
    print("Calls per case: 4 primary element calls + N/A verification only when proposed")
    for i,uid in enumerate(selected,1):
        case=cmap[uid]
        cdir=args.run_root/"cases"/uid
        final_path=cdir/"final_case.json"
        # Fast complete-case resume.
        if final_path.exists():
            old=load_json(final_path)
            if old.get("version")==VERSION and len(old.get("verdicts") or {})==41:
                print(f"[{i}/{len(selected)}] {uid}: COMPLETE cache")
                continue
        staged=base.stage_case(case=case,case_root=case_root,run_dir=args.run_root)
        chunks,cache_hit=base.parse_case_cached(case=case,stage_dir=staged,run_dir=args.run_root,parser_hash=parser_hash)
        chunks=normalize_chunks(chunks)
        print(f"[{i}/{len(selected)}] {uid}: chunks={len(chunks)} {'cache' if cache_hit else 'parsed'}")
        verdicts={}; element_meta={}
        for element in ELEMENTS:
            r,hit=run_element(uid,element,chunks,cp_texts,headings,policy_docs,cdir,args.model)
            print(f"  {element}: {'cache' if hit else 'model'} labels={dict(Counter(x['verdict'] for x in r['verdicts'].values()))}")
            verdicts.update(r["verdicts"]); element_meta[element]={"path":str(cdir/'elements'/f'{element}.json'),"cache":hit}
        if set(verdicts)!={f"CP{x}" for x in range(1,42)}:
            raise RuntimeError(f"{uid}: incomplete CP output {len(verdicts)}")
        bad={cp:r['verdict'] for cp,r in verdicts.items() if r['verdict'] not in ALLOWED}
        if bad: raise RuntimeError(f"{uid}: bad verdicts {bad}")
        final={"schema":"freca-v6.6-deadline-final-case-v1","version":VERSION,"case_uid":uid,"model":args.model,
               "verdicts":verdicts,"element_outputs":element_meta,
               "label_counts":dict(Counter(x['verdict'] for x in verdicts.values()))}
        save_json(final,final_path)
        print("  FINAL",final['label_counts'])
    print("DONE",len(selected),"cases")

if __name__=="__main__":
    main()
