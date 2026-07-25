from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from freca.manifest import sha256_file
from freca.models import (
    Applicability,
    AuditDecision,
    CaseManifest,
    CaseRecord,
    Verdict,
)
from freca.submission import assemble_submission


def _manifest(tmp_path: Path) -> CaseManifest:
    cases = [
        CaseRecord(
            case_id=case_id,
            re_number=(
                "RE-WA-2021-0077"
                if case_id in {35, 100}
                else f"RE-TEST-{case_id:04d}"
            ),
            sources=[],
        )
        for case_id in range(1, 101)
    ]
    return CaseManifest(cases_root=tmp_path, cases=cases, source_count=0)


def _decisions() -> list[AuditDecision]:
    return [
        AuditDecision(
            case_id=case_id,
            cp_id=f"CP{cp}",
            applicability=Applicability.APPLICABLE,
            regulatory_requirement="Requirement from policy.",
            policy_citations=["p1"],
            supporting_evidence=["e1"],
            contrary_evidence=[],
            contradictions=[],
            verdict=Verdict.COMPLIANT,
            reasoning_summary="Supported.",
            confidence=0.9,
            retrieval_complete=True,
        )
        for case_id in range(1, 101)
        for cp in range(1, 42)
    ]


def _template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "All Elements"
    sheet.append(["RE Number"] + [f"CP{cp}" for cp in range(1, 42)])
    workbook.save(path)


def test_submission_rejects_unresolved_tasks(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    _template(template)
    with pytest.raises(ValueError, match="unresolved tasks"):
        assemble_submission(
            _decisions(),
            _manifest(tmp_path),
            template,
            tmp_path / "out.xlsx",
            unresolved_tasks=1,
            allow_unconfirmed_identifiers=True,
        )


def test_submission_blocks_duplicate_re_without_explicit_candidate_override(
    tmp_path: Path,
) -> None:
    template = tmp_path / "template.xlsx"
    _template(template)
    with pytest.raises(ValueError, match="duplicate RE Number"):
        assemble_submission(
            _decisions(), _manifest(tmp_path), template, tmp_path / "out.xlsx"
        )


def test_candidate_submission_has_100_rows_42_columns_and_preserves_template(
    tmp_path: Path,
) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "candidate.xlsx"
    _template(template)
    original_hash = sha256_file(template)

    report = assemble_submission(
        _decisions(),
        _manifest(tmp_path),
        template,
        output,
        allow_unconfirmed_identifiers=True,
    )

    assert sha256_file(template) == original_hash
    workbook = load_workbook(output, data_only=True)
    sheet = workbook.active
    assert sheet.max_row == 101
    assert sheet.max_column == 42
    assert {sheet.cell(row, column).value for row in range(2, 102) for column in range(2, 43)} == {"1"}
    assert report.decision_count == 4100
    assert report.candidate_only is True


def test_submission_rejects_missing_case_cp_decision(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    _template(template)
    with pytest.raises(ValueError, match="expected 4100 decisions"):
        assemble_submission(
            _decisions()[:-1],
            _manifest(tmp_path),
            template,
            tmp_path / "out.xlsx",
            allow_unconfirmed_identifiers=True,
        )
