from __future__ import annotations

import shutil
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from freca.manifest import sha256_file
from freca.models import AuditDecision, CaseManifest, SubmissionReport


_EXPECTED_HEADER = ["RE Number"] + [f"CP{cp}" for cp in range(1, 42)]
_ALLOWED = {"1", "0", "N/A"}


def _decision_map(decisions: list[AuditDecision]) -> dict[tuple[int, str], AuditDecision]:
    if len(decisions) != 4100:
        raise ValueError(f"expected 4100 decisions, got {len(decisions)}")
    mapping: dict[tuple[int, str], AuditDecision] = {}
    for decision in decisions:
        key = (decision.case_id, decision.cp_id)
        if key in mapping:
            raise ValueError(f"duplicate decision: case {decision.case_id} {decision.cp_id}")
        mapping[key] = decision
    expected = {
        (case_id, f"CP{cp}") for case_id in range(1, 101) for cp in range(1, 42)
    }
    if set(mapping) != expected:
        missing = sorted(expected - set(mapping))[:10]
        unexpected = sorted(set(mapping) - expected)[:10]
        raise ValueError(f"decision key mismatch: missing={missing}, unexpected={unexpected}")
    return mapping


def assemble_submission(
    decisions: list[AuditDecision],
    manifest: CaseManifest,
    template_path: Path,
    output_path: Path,
    *,
    unresolved_tasks: int = 0,
    allow_unconfirmed_identifiers: bool = False,
) -> SubmissionReport:
    if unresolved_tasks:
        raise ValueError(f"cannot assemble with {unresolved_tasks} unresolved tasks")
    if output_path.resolve() == template_path.resolve():
        raise ValueError("output must not overwrite the official template")
    mapping = _decision_map(decisions)
    if len(manifest.cases) != 100:
        raise ValueError(f"expected 100 manifest cases, got {len(manifest.cases)}")
    re_counts = Counter(case.re_number for case in manifest.cases)
    duplicates = sorted(re_number for re_number, count in re_counts.items() if count > 1)
    if duplicates and not allow_unconfirmed_identifiers:
        raise ValueError(
            "duplicate RE Number requires organizer confirmation or explicit candidate override: "
            + ", ".join(duplicates)
        )

    template_workbook = load_workbook(template_path, data_only=False)
    template_sheet = template_workbook.active
    header = [template_sheet.cell(1, column).value for column in range(1, 43)]
    if header != _EXPECTED_HEADER:
        raise ValueError("submission template header does not match RE Number, CP1..CP41")
    if template_sheet.max_row != 1:
        raise ValueError(
            "this assembler expects the verified header-only template; a populated official "
            "template requires a separate identifier-order review"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook.active
    for case in sorted(manifest.cases, key=lambda item: item.case_id):
        row = [case.re_number]
        for cp in range(1, 42):
            verdict = mapping[(case.case_id, f"CP{cp}")].verdict.value
            if verdict not in _ALLOWED:
                raise ValueError(f"invalid verdict {verdict}")
            row.append(verdict)
        sheet.append(row)
    workbook.save(output_path)

    reloaded = load_workbook(output_path, data_only=True)
    output_sheet = reloaded.active
    if output_sheet.max_row != 101 or output_sheet.max_column != 42:
        raise ValueError(
            f"output shape is {output_sheet.max_row}x{output_sheet.max_column}, expected 101x42"
        )
    output_header = [output_sheet.cell(1, column).value for column in range(1, 43)]
    if output_header != _EXPECTED_HEADER:
        raise ValueError("output header changed during assembly")
    values = [
        output_sheet.cell(row, column).value
        for row in range(2, 102)
        for column in range(2, 43)
    ]
    if any(value not in _ALLOWED for value in values):
        raise ValueError("output contains blank or invalid verdict cells")
    return SubmissionReport(
        output_path=output_path.resolve(),
        rows=100,
        columns=42,
        decision_count=len(values),
        candidate_only=bool(duplicates) or template_sheet.max_row == 1,
        duplicate_re_numbers=duplicates,
        sha256=sha256_file(output_path),
    )
