"""Curated scoring-criteria table — the team's merged rubric asset.

The xlsx holds one row per checking point with the human-curated red line
and the full scoring standard (threshold materials merged in). It is an
alternative *input* to Stage B: the rubric generator may receive each
row as a pseudo policy chunk so the curated standard — not just clauses
retrieved from the raw Rules PDF — shapes the rubric.

Anti-hardcoding discipline (cf. ``freca.ledger.rubric``):

* this module never references a specific CP id, case, or verdict;
* the Act-layer reference column is intentionally never read into model
  content, matching the team convention that Act material is for human
  traceability only;
* the file is validated structurally (header, CP completeness), not by
  content.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from freca.manifest import sha256_file
from freca.models import ContentKind, EvidenceChunk, SourceLocation, SourceType

CURATED_CHUNK_PREFIX = "curated:"
CURATED_SOURCE_ID = "curated-criteria"
CURATED_PARSER_NAME = "freca.ledger.criteria"
CURATED_PARSER_VERSION = "1"

_EXPECTED_HEADER = (
    "CP",
    "CP定义(中)",
    "红线R3(中)",
    "评分标准（最终版·含门槛/依据材料）",
    "Act层参考(联网核验·非本地材料)",
    "来源说明",
)
_EXPECTED_CP_COUNT = 41


@dataclass(frozen=True)
class CriteriaEntry:
    redline: str
    criteria_text: str
    row_index: int


@dataclass(frozen=True)
class CriteriaTable:
    entries: dict[str, CriteriaEntry]
    sha256: str
    source_name: str
    sheet_name: str

    def entry(self, cp_id: str) -> CriteriaEntry:
        try:
            return self.entries[cp_id]
        except KeyError:
            raise KeyError(f"curated criteria xlsx has no row for {cp_id}") from None

    @classmethod
    def load(cls, path: Path) -> "CriteriaTable":
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            sheet = workbook.worksheets[0]
            sheet_title = sheet.title
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None or tuple(
                "" if cell is None else str(cell).strip() for cell in header
            ) != _EXPECTED_HEADER:
                raise ValueError(
                    f"curated criteria xlsx header mismatch in {path.name}: {header!r}"
                )
            entries: dict[str, CriteriaEntry] = {}
            for index, row in enumerate(rows, start=2):
                cp_id = "" if row[0] is None else str(row[0]).strip()
                if not cp_id:
                    continue
                if cp_id in entries:
                    raise ValueError(f"duplicate curated criteria row: {cp_id}")
                redline = "" if row[2] is None else str(row[2]).strip()
                criteria_text = "" if row[3] is None else str(row[3]).strip()
                if not redline or not criteria_text:
                    raise ValueError(
                        f"curated criteria row {cp_id} has an empty red line or standard"
                    )
                entries[cp_id] = CriteriaEntry(
                    redline=redline,
                    criteria_text=criteria_text,
                    row_index=index,
                )
        finally:
            workbook.close()
        if len(entries) != _EXPECTED_CP_COUNT:
            missing = _EXPECTED_CP_COUNT - len(entries)
            raise ValueError(
                f"curated criteria xlsx must cover {_EXPECTED_CP_COUNT} checking points; "
                f"{path.name} has {len(entries)} (missing {missing})"
            )
        return cls(
            entries=entries,
            sha256=sha256_file(path),
            source_name=path.name,
            sheet_name=sheet_title,
        )


def curated_chunk(entry: CriteriaEntry, *, cp_id: str, table: CriteriaTable) -> EvidenceChunk:
    """Return the curated standard as one evidence chunk for Stage B."""

    content = (
        f"红线R3（判定命题）:\n{entry.redline}\n\n"
        f"评分标准（最终版·含门槛/依据材料）:\n{entry.criteria_text}"
    )
    return EvidenceChunk(
        chunk_id=f"{CURATED_CHUNK_PREFIX}{cp_id}",
        source_id=CURATED_SOURCE_ID,
        source_file=table.source_name,
        source_type=SourceType.XLSX,
        location=SourceLocation(sheet=table.sheet_name),
        content=content,
        content_kind=ContentKind.PARAGRAPH,
        parser_name=CURATED_PARSER_NAME,
        parser_version=CURATED_PARSER_VERSION,
        source_sha256=table.sha256,
        flags=["curated"],
        metadata={"row_index": entry.row_index},
    )


__all__ = [
    "CURATED_CHUNK_PREFIX",
    "CURATED_SOURCE_ID",
    "CriteriaEntry",
    "CriteriaTable",
    "curated_chunk",
]
