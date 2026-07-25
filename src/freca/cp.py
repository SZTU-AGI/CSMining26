from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from freca.manifest import sha256_file
from freca.models import CheckpointDefinition, SourceRecord, SourceType


def load_checkpoints(path: Path) -> list[CheckpointDefinition]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet = workbook.active
    current_element = ""
    current_section = ""
    checkpoints: list[CheckpointDefinition] = []
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(1, column).value:
            current_element = str(sheet.cell(1, column).value).strip()
        if sheet.cell(2, column).value:
            current_section = str(sheet.cell(2, column).value).strip()
        cp_id = str(sheet.cell(4, column).value or "").strip()
        text = str(sheet.cell(3, column).value or "").strip()
        match = re.fullmatch(r"Element-(\d)", current_element, re.IGNORECASE)
        if not cp_id or not text or match is None:
            raise ValueError(f"invalid checkpoint column {column}")
        checkpoints.append(
            CheckpointDefinition(
                cp_id=cp_id,
                element_id=int(match.group(1)),
                element_title=current_element,
                section_title=current_section,
                text=text,
                source_file=path.name,
                cell=sheet.cell(3, column).coordinate,
            )
        )
    expected = [f"CP{index}" for index in range(1, 42)]
    actual = [checkpoint.cp_id for checkpoint in checkpoints]
    if actual != expected:
        raise ValueError(f"expected CP1..CP41, got {actual}")
    return checkpoints


def build_policy_source(path: Path) -> SourceRecord:
    path = path.resolve()
    return SourceRecord(
        source_id="policy-rules-2021",
        path=path,
        source_type=SourceType.PDF,
        sha256=sha256_file(path),
    )
