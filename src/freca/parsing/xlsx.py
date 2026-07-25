from __future__ import annotations

import re
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from freca.models import ContentKind, EvidenceChunk, SourceLocation, SourceRecord
from freca.parsing.chunking import stable_chunk_id


_RE_NUMBER_PATTERN = re.compile(r"\bRE-[A-Z]{2,3}-\d{4}-\d{4}\b", re.IGNORECASE)


def _format_value(value: object) -> str:
    if value is None:
        return "<BLANK>"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def parse_xlsx(source: SourceRecord, *, max_rows: int = 20) -> list[EvidenceChunk]:
    if max_rows < 1:
        raise ValueError("max_rows must be positive")
    workbook = load_workbook(source.path, data_only=False, read_only=False)
    chunks: list[EvidenceChunk] = []
    for sheet in workbook.worksheets:
        if sheet.max_row < 1 or sheet.max_column < 1:
            continue
        for start_row in range(1, sheet.max_row + 1, max_rows):
            end_row = min(sheet.max_row, start_row + max_rows - 1)
            start_cell = f"A{start_row}"
            end_cell = f"{get_column_letter(sheet.max_column)}{end_row}"
            cell_range = f"{start_cell}:{end_cell}"
            lines: list[str] = []
            formulas: list[str] = []
            for row in sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=1,
                max_col=sheet.max_column,
            ):
                parts = []
                for cell in row:
                    parts.append(f"{cell.coordinate}={_format_value(cell.value)}")
                    if cell.data_type == "f":
                        formulas.append(cell.coordinate)
                lines.append(" | ".join(parts))
            locator = f"sheet-{sheet.title}-cells-{cell_range}"
            content = "\n".join(lines)
            embedded_re_numbers = sorted(
                {match.upper() for match in _RE_NUMBER_PATTERN.findall(content)}
            )
            flags = list(source.flags)
            if source.re_number and any(
                value != source.re_number.upper() for value in embedded_re_numbers
            ):
                flags.append("embedded_re_number_mismatch")
            chunks.append(
                EvidenceChunk(
                    chunk_id=stable_chunk_id(source, locator),
                    case_id=source.case_id,
                    re_number=source.re_number,
                    track=source.track,
                    source_id=source.source_id,
                    source_file=source.path.name,
                    source_type=source.source_type,
                    location=SourceLocation(sheet=sheet.title, cell_range=cell_range),
                    content=content,
                    content_kind=ContentKind.TABLE,
                    parser_name="openpyxl",
                    parser_version="1",
                    source_sha256=source.sha256,
                    flags=flags,
                    metadata={
                        "formula_cells": formulas,
                        "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                        "embedded_re_numbers": embedded_re_numbers,
                    },
                )
            )
    return chunks
